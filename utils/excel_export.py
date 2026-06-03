"""将 OCR 结果导出为 xlsx。"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config as cfg


def _safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r'[\\/*?:\[\]]', "_", name)
    base = base[:28] or "Sheet"
    candidate = base
    n = 1
    while candidate in used:
        suffix = f"_{n}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def _autosize_columns(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[letter].width = max_len + 2


def export_results_to_xlsx(
    results: list[dict],
    *,
    output_dir: str | None = None,
) -> str:
    """
    results: [{"filename": "a.jpg", "headers": [...], "rows": [[...]]}, ...]
    返回生成的 xlsx 绝对路径。
    """
    out_dir = Path(output_dir or cfg.EXPORTS_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"ocr_result_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    for item in results:
        filename = item.get("filename") or "unknown"
        headers = item.get("headers") or []
        rows = item.get("rows") or []
        sheet_name = _safe_sheet_name(Path(filename).stem, used_names)
        ws = wb.create_sheet(title=sheet_name)

        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            if isinstance(row, list):
                ws.append(row)
            elif isinstance(row, dict) and headers:
                ws.append([row.get(h, "") for h in headers])

        col_count = max(len(headers), 1)
        if ws.max_row:
            _autosize_columns(ws, col_count)

    if not wb.sheetnames:
        ws = wb.create_sheet("空结果")
        ws.append(["无识别数据"])

    wb.save(out_path)
    return str(out_path.resolve())
